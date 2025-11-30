import os
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side


class WorkSheetStyle:

    def __init__(self):
        self.width_of_columns = {15: ('B', 'C'), 5: ('A',)}
        self.height_of_rows = {70: (4, 7), 35: (5, 14), 25: (*range(16, 24), *range(28, 36), 26)}
        self.merge_columns_of_row = {
            'C:D': '4', 'E:H': '4', 'C:H': '5:6:7:8:9:10:11:12', 'E:K': '14', 'A:C': '24:25:26:27', 'B:H': '2',
        }
        self.merge_rows_of_column = {'14:15': 'A:B:C:D'}
        self.merge_rows_and_columns: list[str] = []
        self.border_cells = {
            self.__create_border(side_style='thin'): ('E16:K23',),
            self.__create_border(): ('B4:I12', 'A14:D35', 'E14:K15', 'E24:K28'),
        }

    @property
    def cells_to_merge(self) -> tuple[str, ...]:
        return tuple(
            self.__get_range_cells_from_rows()
            + self.__get_range_cells_from_columns()
            + self.merge_rows_and_columns
        )

    def __get_range_cells_from_columns(self) -> list[str]:
        range_of_cells = []
        for columns, rows in self.merge_columns_of_row.items():
            column_start, column_end = columns.split(':')
            for row in rows.split(':'):
                 range_of_cells.append(':'.join((column_start + row, column_end + row)))
        return range_of_cells

    def __get_range_cells_from_rows(self) -> list[str]:
        range_of_cells = []
        for rows, columns in self.merge_rows_of_column.items():
            row_start, row_end = rows.split(':')
            for column in columns.split(':'):
                 range_of_cells.append(':'.join((column + row_start, column + row_end)))
        return range_of_cells

    @staticmethod
    def __create_border(side_style='medium', border_type='full') -> Border:
        type_side = Side(style=side_style)
        match border_type.lower():
            case 'top':
                return Border(top=type_side)
            case _:
                return Border(top=type_side, right=type_side, bottom=type_side, left=type_side)


class WorkSheet:

    def __init__(self, wb_title: str, ws_title: str) -> None:
        self.ws_style = WorkSheetStyle()
        self.wb_title = wb_title
        self.ws_title = ws_title
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_style = WorkSheetStyle()

    def change_ws_title(self) -> None:
        self.ws.title = self.ws_title
        self.__set_rows_height(self.ws_style)
        self.__set_columns_width(self.ws_style)
        self.__merge_cells(self.ws_style.cells_to_merge)
        self.__frame_for_cells()

    def save_work_sheet(self):
        self.wb.save(self.wb_title)

    def __set_columns_width(self, ws_style: WorkSheetStyle) -> None:
        for width, columns in ws_style.width_of_columns.items():
            for column in columns:
                self.ws.column_dimensions[column].width = width

    def __set_rows_height(self, ws_style: WorkSheetStyle) -> None:
        for height, rows in ws_style.height_of_rows.items():
            for row in rows:
                self.ws.row_dimensions[row].height = height

    def __merge_cells(self, range_cells: tuple[str, ...]) -> None:
        for cells in range_cells:
            self.ws.merge_cells(cells)

    def __frame_for_cells(self) -> None:
        for border, cells_range in self.ws_style.border_cells.items():
            for row_range in cells_range:
                for row in self.ws[row_range]:
                    for cell in row:
                        cell.border = border


if __name__ == '__main__':
    b_title = 'Пробная_ведомость_работ.xlsx'
    ws_first_title = 'Тестовая страница V05110______'
    if os.path.exists(b_title):
        os.remove(b_title)
    work_sheet = WorkSheet(b_title, ws_first_title)
    work_sheet.change_ws_title()
    work_sheet.save_work_sheet()

